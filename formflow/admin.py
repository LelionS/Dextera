from django.contrib import admin
from django.apps import apps
from django.contrib.auth.models import User
from django.contrib.auth.admin import UserAdmin
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
import openpyxl
from django import forms
from io import TextIOWrapper
import csv

from .models import House, Bay, Bed, Variety

class VarietyImportForm(forms.Form):
    excel_file = forms.FileField()

def import_variety_view(request):
    if request.method == "POST":
        form = VarietyImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["excel_file"]
            try:
                wb = openpyxl.load_workbook(excel_file)
                sheet = wb.active

                # Assume headers: house | bay | bed | name
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    house_name, bay_name, bed_code, variety_name = row
                    if not all([house_name, bay_name, bed_code, variety_name]):
                        continue  # Skip incomplete rows

                    house, _ = House.objects.get_or_create(name=house_name)
                    bay, _ = Bay.objects.get_or_create(name=bay_name, house=house)
                    bed, _ = Bed.objects.get_or_create(code=bed_code, bay=bay)
                    Variety.objects.create(name=variety_name, bed=bed)

                messages.success(request, "Varieties imported successfully from Excel.")
                return redirect("..")
            except Exception as e:
                messages.error(request, f"Error reading Excel file: {e}")
    else:
        form = VarietyImportForm()
    return render(request, "admin/import_variety.html", {"form": form})

# Deactivate user action
def deactivate_users(modeladmin, request, queryset):
    queryset.update(is_active=False)
deactivate_users.short_description = "Deactivate selected users"

# Custom user admin
class CustomUserAdmin(UserAdmin):
    actions = [deactivate_users]

    class Media:
        css = {
            'all': (
                'https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css',
                '/static/admin/css/custom_admin.css',
            )
        }
        js = ('/static/admin/js/custom_admin.js',)

admin.site.unregister(User)
admin.site.register(User, CustomUserAdmin)

# Inlines
class VarietyInline(admin.TabularInline):
    model = Variety
    extra = 1

class BedInline(admin.TabularInline):
    model = Bed
    extra = 1
    show_change_link = True

class BayInline(admin.TabularInline):
    model = Bay
    extra = 1
    show_change_link = True

# Admin registrations
@admin.register(House)
class HouseAdmin(admin.ModelAdmin):
    list_display = ("name",)
    search_fields = ("name",)
    inlines = [BayInline]

@admin.register(Bay)
class BayAdmin(admin.ModelAdmin):
    list_display = ("name", "house")
    search_fields = ("name",)
    list_filter = ("house",)
    inlines = [BedInline]

@admin.register(Bed)
class BedAdmin(admin.ModelAdmin):
    list_display = ("code", "bay")
    search_fields = ("code",)
    list_filter = ("bay",)
    inlines = [VarietyInline]

import openpyxl
from django.http import HttpResponse

@admin.register(Variety)
class VarietyAdmin(admin.ModelAdmin):
    change_list_template = "admin/variety_change_list.html"
    list_display = ("name", "get_house","get_bay", "bed", "location")
    search_fields = ("name",)
    list_filter = ("bed",)
    actions = ["export_to_excel"]

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("import-varieties/", self.admin_site.admin_view(import_variety_view), name="import-varieties"),
        ]
        return custom_urls + urls

    def get_bay(self, obj):
        return obj.bed.bay.name
    get_bay.short_description = "Bay"

    def get_house(self, obj):
        return obj.bed.bay.house.name
    get_house.short_description = "House"

    def location(self, obj):
        return f"{obj.bed.bay.house.name} - {obj.bed.bay.name} - {obj.bed.code}"
    location.short_description = "Location"

    def export_to_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Varieties"

        # Headers
        ws.append(["Name", "Bed", "Bay", "House", "Location"])

        # Data rows
        for variety in queryset:
            bed = variety.bed
            bay = bed.bay
            house = bay.house
            location = f"{house.name} > {bay.name} > {bed.code}"
            ws.append([variety.name, bed.code, bay.name, house.name, location])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="varieties.xlsx"'
        wb.save(response)
        return response
    export_to_excel.short_description = "Export selected varieties to Excel"


# Register remaining models automatically
explicit_models = {House, Bay, Bed, Variety}
app_models = apps.get_app_config('formflow').get_models()
for model in app_models:
    if model not in explicit_models and model not in admin.site._registry:
        admin.site.register(model)
